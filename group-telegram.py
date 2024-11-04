import os
from flask import Flask, send_from_directory
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
from openpyxl import Workbook, load_workbook
from datetime import datetime
import threading

# إعداد القروبات
BAHRAIN_GROUP_LINK = "@Rashed_bahrain"
OTHER_GROUP_LINK = "@Rashed_GCC"

# إعداد مسار ملف Excel
FILE_PATH = "responses.xlsx"

# الأسئلة
QUESTIONS = [
    ("شنو جنسيتك؟", ["بحريني", "سعودي", "إماراتي", "كويتي", "قطري", "عُماني", "دول أخرى"]),
    ("كم عمرك؟", ["10-15 سنة", "16-20 سنة", "21-35 سنة", "36-46 سنة", "47+"]),
    ("شنو جنسك؟", ["ذكر", "أنثى"]),
    ("هل أنت مقيم في البحرين؟", ["نعم", "لا"])
]

# إنشاء أو تحديث ملف Excel
def save_to_excel(username, answers, phone_number=None):
    file_exists = os.path.isfile(FILE_PATH)
    if file_exists:
        workbook = load_workbook(FILE_PATH)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        headers = ["تاريخ الإضافة", "اسم المستخدم", "الجنسية", "العمر", "الجنس", "الإقامة في البحرين", "رقم الهاتف"]
        sheet.append(headers)
    
    add_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [add_date, username, *answers, phone_number if phone_number else ""]
    sheet.append(row)
    workbook.save(FILE_PATH)
    print(f"Data saved to {FILE_PATH}")

# رسالة ترحيب أولية
async def welcome_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("مرحباً بك في البوت! اضغط /start للبدء في الاستبيان.")

# رسالة البداية
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data["answers"] = []
    await update.message.reply_text("مرحباً! سأطرح عليك سلسلة من الأسئلة لتوجيهك إلى القروب المناسب. أجب على الأسئلة التالية.")
    await ask_question(update, context)

# طرح السؤال التالي
async def ask_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    answers = context.user_data["answers"]
    question, options = QUESTIONS[len(answers)]
    buttons = [[InlineKeyboardButton(option, callback_data=option)] for option in options]
    reply_markup = InlineKeyboardMarkup(buttons)
    await update.message.reply_text(question, reply_markup=reply_markup)

# التعامل مع الإجابات
async def handle_answer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["answers"].append(query.data)

    if len(context.user_data["answers"]) < len(QUESTIONS):
        await ask_question(query, context)
    else:
        await finish_quiz(query, context)

# إنهاء الاستبيان وحفظ البيانات
async def finish_quiz(query, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = query.from_user
    answers = context.user_data["answers"]
    nationality, _, _, resident = answers

    group_link = BAHRAIN_GROUP_LINK if nationality == "بحريني" else OTHER_GROUP_LINK

    phone_number = query.message.contact.phone_number if query.message.contact else None
    save_to_excel(user.username, answers, phone_number)

    await query.message.reply_text(f"شكرًا لإجابتك على الأسئلة! يمكنك الانضمام إلى القروب المناسب من هنا: {group_link}")

# إعداد تطبيق Telegram
def start_telegram_bot():
    app = Application.builder().token("7324354293:AAESUs8cyUVS6lt1TXE3hNVx4uC3u1nBSfU").build()
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, welcome_message))
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(handle_answer))
    app.run_polling()

# إعداد Flask للتنزيل
flask_app = Flask(__name__)

@flask_app.route('/download')
def download_file():
    return send_from_directory(os.getcwd(), FILE_PATH, as_attachment=True)

# بدء Flask وTelegram معاً
if __name__ == "__main__":
    threading.Thread(target=start_telegram_bot).start()
    flask_app.run(host="0.0.0.0", port=5000)
